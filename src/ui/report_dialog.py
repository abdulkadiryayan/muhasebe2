from PySide6.QtWidgets import (QDialog, QVBoxLayout, QFormLayout, QComboBox,
                              QPushButton, QMessageBox)
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

class ReportDialog(QDialog):
    def __init__(self, database):
        super().__init__()
        self.database = database
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Rapor Oluştur")
        self.setMinimumWidth(300)
        layout = QVBoxLayout(self)
        form = QFormLayout()

        # Başlık seçimi
        self.title_combo = QComboBox()
        self.title_combo.addItem("Tümü", None)
        self.refresh_titles()
        form.addRow("Başlık:", self.title_combo)

        # Kasa sahibi seçimi
        self.cash_owner_combo = QComboBox()
        self.cash_owner_combo.addItem("Tümü", None)
        self.refresh_cash_owners()
        form.addRow("Kasa Sahibi:", self.cash_owner_combo)

        layout.addLayout(form)

        # Rapor oluştur butonu
        btn_create = QPushButton("Excel Raporu Oluştur")
        btn_create.clicked.connect(self.create_report)
        layout.addWidget(btn_create)

    def refresh_titles(self):
        titles = self.database.get_titles()
        for title in titles:
            self.title_combo.addItem(title['name'], title['id'])

    def refresh_cash_owners(self):
        cash_owners = self.database.get_cash_owners()
        for owner in cash_owners:
            self.cash_owner_combo.addItem(owner['name'], owner['id'])

    def create_report(self):
        filters = {}
        
        if self.title_combo.currentData():
            filters['title_id'] = self.title_combo.currentData()
        
        if self.cash_owner_combo.currentData():
            filters['cash_owner_id'] = self.cash_owner_combo.currentData()

        transactions = self.database.get_transactions(filters)
        
        if not transactions:
            QMessageBox.warning(self, "Uyarı", "Seçilen kriterlere uygun kayıt bulunamadı!")
            return

        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Muhasebe Raporu"

        # Stil tanımlamaları
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Kenarlık stilleri
        thin_border = Side(border_style="thin", color="000000")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Hücre stilleri
        number_format = '#,##0.00'
        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        # Başlıkları ekle
        headers = ["Tarih", "Başlık", "Kasa Sahibi", "Firma", "Açıklama",
                  "Yapılan Ödeme", "Alınan Ödeme", "Alınan Çek", "Verilen Çek",
                  "Daire Satış", "Fatura Tutarı", "Miktar", "Birim Fiyat",
                  "Toplam Tutar"]

        # Sütun genişlikleri
        column_widths = {
            'A': 15,  # Tarih
            'B': 25,  # Başlık
            'C': 25,  # Kasa Sahibi
            'D': 25,  # Firma
            'E': 40,  # Açıklama
            'F': 15,  # Yapılan Ödeme
            'G': 15,  # Alınan Ödeme
            'H': 15,  # Alınan Çek
            'I': 15,  # Verilen Çek
            'J': 15,  # Daire Satış
            'K': 15,  # Fatura Tutarı
            'L': 12,  # Miktar
            'M': 15,  # Birim Fiyat
            'N': 15,  # Toplam Tutar
        }

        # Sütun genişliklerini ayarla
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Başlıkları formatla
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Satır yüksekliği
        ws.row_dimensions[1].height = 30

        # Verileri ekle ve formatla
        for row, trans in enumerate(transactions, 2):
            # Tarih ve metin verileri
            ws.cell(row=row, column=1, value=trans['date']).alignment = cell_alignment
            ws.cell(row=row, column=2, value=trans['title_name']).alignment = cell_alignment
            ws.cell(row=row, column=3, value=trans['cash_owner_name']).alignment = cell_alignment
            ws.cell(row=row, column=4, value=trans['company_name']).alignment = cell_alignment
            ws.cell(row=row, column=5, value=trans['description']).alignment = cell_alignment

            # Sayısal veriler
            numeric_cells = [
                (6, trans['expense']),
                (7, trans['payment_received']),
                (8, trans['check_received']),
                (9, trans['check_given']),
                (10, trans['apartment_sale']),
                (11, trans['invoice_amount']),
                (12, trans['quantity']),
                (13, trans['unit_price'])
            ]

            for col, value in numeric_cells:
                cell = ws.cell(row=row, column=col, value=value or 0)
                cell.number_format = number_format
                cell.alignment = number_alignment

            # Toplam Tutar hesaplama ve formatlama
            total_cell = ws.cell(row=row, column=14)
            total_cell.value = f"=L{row}*M{row}"
            total_cell.number_format = number_format
            total_cell.alignment = number_alignment

            # Tüm hücrelere kenarlık ekle
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = border

        # Alternatif satır renklendirmesi
        for row in range(2, len(transactions) + 2):
            if row % 2 == 0:
                for col in range(1, 15):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Masaüstüne kaydet
        desktop_path = str(Path.home() / "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = f"{desktop_path}/muhasebe_raporu_{timestamp}.xlsx"
        
        wb.save(excel_path)
        QMessageBox.information(self, "Başarılı", f"Rapor başarıyla oluşturuldu:\n{excel_path}") 